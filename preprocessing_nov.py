try:
    import os
    import pandas as pd
    import openpyxl as opyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    import re
except ImportError as import_err:
    print(import_err)

# Set pandas options
pd.set_option("display.max_columns", None)
pd.options.mode.copy_on_write = True


class PreprocessingDataClass:
    def __init__(self, excel_file: str, columns_to_drop: list[str], output_csv_path: str):
        self.excel_file = excel_file
        self.input_sheet = "input"
        self.output_sheet = "output"
        self.excel_df = pd.read_excel(self.excel_file,
                                      sheet_name=self.input_sheet,
                                      header=0,
                                      dtype=str)
        self.columns_to_drop = columns_to_drop
        self.output_csv_path = output_csv_path
        self.filtered_df = None
        self.address_name = "full_address"
        self.aa = "AA"
        self.columns_order = [
            "AA",
            "aa",
            "source",
            "streetName_prefix",
            "full_address",
            "address",
            "neighborhood",
            "drop",
            "klm",
            "road",
            "po_box",
            "poi",
            "town",
            "municipal_unit",
            "settlement",
            "toponym",
            "city_block",
            "unit",
            "number",
            "check",
            "data_city",
            "data_zip",
            "δρομος",
            "ονομασία οδού ok",
            "Street",
            "Num",
            "City",
            "Zip",
            "Municipality",
            "Prefecture",
            "Type",
            "x",
            "y",
        ]

    def get_column_order(self, df: pd.DataFrame, desired_order: list) -> list:
        """
        Reorders the columns of a DataFrame based on a desired order of prefixes.

        Parameters:
            df (pd.DataFrame): The DataFrame containing columns to reorder.
            desired_order (list): List of column prefixes in the desired order.

        Returns:
            list: List of columns ordered based on the desired prefixes.
        """
        try:
            # Initialize an empty list to hold the final order of columns
            final_order = []

            # Iterate over each prefix in the desired order list
            for column in desired_order:
                # Find columns in df that start with the current prefix and are not already in final_order
                matching_columns = [c for c in df.columns if c.startswith(column) and c not in final_order]
                # Add these matching columns to the final_order list
                final_order.extend(matching_columns)

            # Return the list of columns in the final order
            return final_order
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def drop_columns(self) -> None:
        """
        Drops specific columns from the DataFrame based on tags or a list of columns to drop.
        """
        try:
            # Check if any column is tagged as "drop_field"
            drop_fields = [column for column in self.excel_df.columns if "drop_field" in column]

            # Drop the columns that are tagged as "drop_field"
            self.excel_df = self.excel_df.drop(
                columns=self.excel_df.columns[self.excel_df.columns.map(lambda x: x in drop_fields)]
            )

            # Check if there are any columns specified to be dropped via the "columns_to_drop" attribute
            if self.columns_to_drop:
                # Verify that all columns to drop exist in the DataFrame
                if set(self.columns_to_drop) <= set(self.excel_df.columns.to_list()):
                    # Print the sorted list of columns that will be deleted
                    print(f"Deleting columns -> {sorted(self.columns_to_drop)}")

                    # Drop the columns specified in "columns_to_drop"
                    self.excel_df = self.excel_df.drop(
                        columns=self.excel_df.columns[
                            self.excel_df.columns.map(lambda x: x.split('.')[0] in self.columns_to_drop)]
                    )
                else:
                    # Print an error message if some columns in "columns_to_drop" do not exist in the DataFrame
                    print(
                        f"Input columns to drop: {set(self.columns_to_drop) - set(self.excel_df.columns.to_list())} are not valid."
                    )
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def remove_quotes(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Removes double and single quotes from all string columns in the given DataFrame.

        Args:
        - df (pd.DataFrame): The input pandas DataFrame from which quotes will be removed.

        Returns:
        - pd.DataFrame: A new DataFrame with quotes removed from all string columns.
        """
        try:
            # Select columns with object data type (typically string columns)
            str_cols = df.select_dtypes(include=["object"]).columns

            # Apply the lambda function to remove double quotes (") and single quotes (')
            df[str_cols] = df[str_cols].apply(
                lambda col: col.str.replace("\"", "").str.replace("'", "")
            )

            # Return the modified DataFrame
            return df
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def build_rest(self) -> None:
        """
        This method constructs a 'rest' column in the DataFrame, based on the values of an address-related column.
        It iteratively checks columns (defined in `columns_to_merge`), and if any values from these columns are
        present in the 'rest' column, they are removed. The purpose is to keep only the unique portions of the
        address that don't overlap with predefined columns like 'neighborhood', 'town', etc.

        The final 'rest' column is then written back to the original Excel file, inserted in the correct position
        after the 'type' column or replaced if it already exists.
        """
        try:
            # Define a list of column names that are related to address data, and whose values should be removed from 'rest'
            columns_to_merge = [
                "neighborhood",
                "town",
                "municipal_unit",
                "settlement",
                "toponym",
                "city_block",
                "unit",
                "poi",
                "po_box",
                "streetName_prefix",
                "klm",
                "road",
                "data_city",
                "Street",
                "number",
                "check",
                "Municipality",
                "data_zip",
                "Prefecture",
                "Type",
            ]

            # Define the name of the new column to be created
            rest_column = "rest"
            full_address_column = "full_address"
            # type_column = "type"

            # Get only the relevant columns from the DataFrame that match the `columns_to_merge` list
            # This handles cases where columns may have suffixes like ".1", ".2", etc.
            columns_to_merge_num = [column for column in self.excel_df if column.split(".")[0] in columns_to_merge]

            # Initialize the 'rest' column with the values from the main address column
            # This creates a copy of the address column to manipulate later
            self.excel_df[rest_column] = self.excel_df.apply(lambda x: x[self.address_name], axis=1)

            # Iterate over each relevant column and remove its value from the 'rest' column
            for column in columns_to_merge_num:
                # Check each cell, and if the value of the current column is found in the 'rest' column, remove it
                # Use `pd.notna(x[column])` to ensure non-null values are processed
                self.excel_df[rest_column] = self.excel_df.apply(
                    lambda x: x[rest_column].replace(str(x[column]), '') if pd.notna(x[column]) else x[rest_column], axis=1
                )

            # Extract the newly created 'rest' column data for writing back to the Excel sheet
            new_column_data = self.excel_df[rest_column]

            # Open the existing Excel file and select the sheet
            wb = opyxl.load_workbook(self.excel_file)
            ws = wb[self.input_sheet]

            # Find the index of the 'full_address' column to insert the 'rest' column directly after it
            full_address_index = list(map(lambda x: x.lower(), list(self.excel_df.columns))).index(full_address_column)
            insert_position = full_address_index + 1  # Add 1 to get the correct 1-based position in Excel

            # Get the header row from the existing Excel sheet to check if 'rest' already exists
            header_row = [cell.value for cell in ws[1]]  # Assume the first row contains headers

            # If 'rest' column already exists in the sheet, remove it before re-adding
            if rest_column in header_row:
                rest_column_position = header_row.index(rest_column)
                ws.delete_cols(rest_column_position + 1)  # Use 1-based index for deleting the column
                insert_position = rest_column_position  # Reuse the position of the deleted column

            # Insert the new 'rest' column in the desired position
            ws.insert_cols(insert_position + 1)  # Shift columns to the right to make space
            ws.cell(row=1, column=insert_position + 1, value=rest_column)  # Write the header 'rest'

            # Write the 'rest' column data (starting from the second row for actual data)
            for i, value in enumerate(new_column_data, start=2):
                ws.cell(row=i, column=insert_position + 1, value=value)

            # Save the modified Excel workbook
            wb.save(self.excel_file)
        except Exception as main_error:
            # Print any error encountered during the execution of the method
            print(main_error)

    def add_line_numbers(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Adds a new column 'line_number' to the dataframe where it assigns
        the index of each substring in the 'address' column for a specific 'old_index'.

        The columns 'aa', 'address', 'street', 'city', 'zip', 'type' receive a line number of 0.

        Parameters:
        df : pd.DataFrame
            Input dataframe that contains columns like 'old_index', 'target', 'text', etc.

        Returns:
        pd.DataFrame
            Updated dataframe with a new 'line_number' column.
        """
        try:
            # Step 1: Define a list of targets to be excluded from receiving a line number
            exclude_list = [
                "aa", "address", "street", "num", "city", "zip", "type",
                "AA", "Street", "Num", "City", "Zip", "Type"
            ]

            # Step 2: Initialize an empty list to store calculated line numbers
            line_numbers = []

            # Step 3: Group the dataframe by 'old_index' and iterate through each group
            for old_index, group in df.groupby("old_index"):

                # Step 4: Get the row where 'target' is 'full_address' to extract the address components
                address_row = group[group["target"] == "full_address"]

                # If there is a non-empty address row, break and index the address into components
                if not address_row.empty:
                    address = address_row["text"].values[0]

                    # Call method to break the address into components and index them
                    address_components = self.break_and_index_address(address)
                else:
                    # If no address is found, set address_components as an empty dictionary
                    address_components = {}

                # Step 5: Iterate over each row in the group
                for _, row in group.iterrows():
                    # If 'target' is in exclude_list, set line number to 0
                    if row["target"] in exclude_list:
                        line_numbers.append(0)
                    else:
                        # Otherwise, assign line number based on the address components
                        line_number = address_components.get(row["text"], 0)
                        line_numbers.append(line_number)

            # Step 6: Assign the generated line numbers to the 'line_number' column in the dataframe
            df["line_number"] = line_numbers

            # Return the updated dataframe
            return df
        except Exception as error:
            # Print any error encountered during the execution of the method
            print(error)

    def transform_address(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Transforms the address column in the DataFrame by removing specific substrings found in other columns.

        Parameters:
            df (pd.DataFrame): The input DataFrame containing an address column and several other columns
                               with substrings that may need to be removed from the address.

        Returns:
            pd.DataFrame: The DataFrame with a transformed address column where specific substrings have been removed.
        """
        try:
            # Define the name of the address column to be transformed
            address_column = self.address_name

            # Define a new column that will store the modified address
            new_address = "address_to_explode"

            # Define a temporary column to hold combined substrings to be removed from the address
            substrings_combined = "substrings_combined"

            # Copy the original address to the new address column
            df[new_address] = df[address_column]

            # Define a tuple of substring column names that need to be removed from the address
            exclude_substrings_column_tuple = (
                "neighborhood",
                "town",
                "municipal_unit",
                "settlement",
                "toponym",
                "city_block",
                "unit",
                "poi",
                "po_box",
                "streetName_prefix",
                "klm",
                "check",
                "Municipality",
                "data_city",
                "data_zip",
                "Prefecture",
                "Type",
                "drop",
            )

            # Iterate over each substring column in the tuple
            for substring_column in exclude_substrings_column_tuple:
                # Find columns in df that match the current substring column name
                matching_columns = [col for col in df.columns if substring_column in col]

                # If matching columns are found, combine their values into a single string per row
                if matching_columns:
                    df[substrings_combined] = df[matching_columns].apply(
                        lambda row: " ".join(row.dropna().astype(str).str.strip()), axis=1
                    )
                else:
                    df[substrings_combined] = ''
                    continue

                # Define a function to remove substrings from the address column
                def remove_substrings_from_address(row):
                    # Get the address from the new address column
                    address = row[new_address].replace('\"', '')

                    # Get the combined substrings to remove
                    substr_to_remove = row[substrings_combined]

                    # If the substring to remove is valid and exists in the address, remove it
                    if pd.notna(substr_to_remove) and substr_to_remove in address:
                        address = address.replace(substr_to_remove, '').strip()

                    # Return the cleaned address
                    return address

                # Apply the remove_substrings_from_address function to each row in the DataFrame
                df[new_address] = df.apply(remove_substrings_from_address, axis=1)

            # Drop the temporary substrings_combined column as it's no longer needed
            df = df.drop(substrings_combined, axis=1)

            # Return the transformed DataFrame
            return df
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def break_address(self, address: str) -> list[str]:
        """
        Breaks down an address string into a list of substrings, removing commas and empty strings.

        Parameters:
            address (str): The input address string that needs to be split into substrings.

        Returns:
            list[str]: A list of substrings obtained from the address, with commas removed and empty strings excluded.
        """
        try:
            # Replace commas in the address with an empty string, then split the address into substrings
            break_address = [substring for substring in address.replace(r',', '').split() if not substring == '']

            # Return the list of non-empty substrings
            return break_address
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def break_and_index_address(self, address: str) -> dict:
        """
        This function takes a string (e.g., an address), breaks it down into components,
        and assigns an index to each substring. The indexing starts at 1 and each unique
        part of the string is indexed in order of appearance.

        Parameters:
            address (str): The input string (e.g., an address).

        Returns:
            dict: A dictionary where keys are the substrings and values are their indices.
        """
        try:
            # Step 1: Define a regex pattern to match different parts of the address
            # The pattern matches:
            #   - Words with hyphens (e.g., 'word-word')
            #   - Words with periods (e.g., 'word.word.')
            #   - Regular words
            #   - Non-word characters (excluding spaces)
            # pattern = r"\b\w\.\b|\b\w+\.\b|\(\w+(?:['`]\w+)?\b|\b\w+\)|\b\w+(?:['`]\w+)?\-\w+|\w+(?:['`]\w+)?\.(?=\s)|\w+\.\w+\.?|\w+(?:['`]\w+)?\/\w+|\w+(?:['`]\w+)?|[^\w\s]"
            pattern = r'" ?\b\w+ ?"(?!\s)|\' ?\b\w+ ?\'(?!\s)|\b\w\.\b|\b\w+\.\b|\(\w+(?:[\'`]\w+)?\b|\b\w+\)|\b\w+(?:[\'`]\w+)?\-\w+|\w+(?:[\'`]\w+)?\.(?=\s)|\w+\.\w+\.?|\w+(?:[\'`]\w+)?\/\w+|\w+(?:[\'`]\w+)?|[^\w\s]'

            # Step 2: Use re.findall to extract components from the address based on the regex pattern
            components = re.findall(pattern, address, re.UNICODE)

            # Step 3: Filter out unwanted characters like commas, periods, and double quotes
            components = [component for component in components if component not in [",", ".", "\"", "\'", "`"]]

            # Step 4: Assign an index to each unique component starting from 1
            indexed_components = {component: idx for idx, component in enumerate(components, 1)}

            # Return the dictionary containing substrings as keys and their indices as values
            return indexed_components
        except Exception as error:
            # Print any error encountered during the execution of the method
            print(error)

    def explode_address(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Splits address strings in the DataFrame into individual components and expands them into separate rows.

        Parameters:
            df (pd.DataFrame): The input DataFrame containing a "text" column that may contain address strings.

        Returns:
            pd.DataFrame: The DataFrame with address strings split into individual components and exploded into separate rows.
        """
        try:
            # Apply the break_address function to the "text" column if the "target" column is "address_to_explode"
            df["text"] = df.apply(
                lambda row: self.break_address(row["text"]) if row["target"] == "address_to_explode" else row["text"], axis=1
            )

            df = df.explode("text")
            return df
        except Exception as error:
            # Print and handle any error that occurs
            print(error)

    def rename_address_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Renames the 'target' column in the DataFrame based on the content of the 'text' column,
        distinguishing between street numbers and street names.

        Parameters:
            df (pd.DataFrame): The input DataFrame containing a 'target' column that identifies
                               address components and a 'text' column with the address data.

        Returns:
            pd.DataFrame: The DataFrame with updated 'target' values for rows where the 'target' was "address_to_explode".
        """
        try:
            # Reset the DataFrame's index and drop the old index column
            df = df.reset_index().drop(columns=["index"])
            df.fillna(value={"text": " "}, inplace=True)

            # Iterate over each row in the DataFrame
            for idx, row in df.iterrows():
                # Check if the 'target' column is "address_to_explode"
                if row["target"] == "address_to_explode":
                    # Check if 'text' is a number or contains minor non-numeric content like spaces or hyphens
                    if re.fullmatch(
                            r"\d+(?:-\d+)?(?:[Α-ΩA-Za-z](?:'\s*[\dΑ-ΩA-Za-z]*)*)?|\d+(?!ης)[α-ωΑ-Ω]+(?:[-\s]*\d+[α-ωΑ-ΩA-Za-z]*)*",
                            row["text"]):
                        # If the text matches a street number pattern, set 'target' to "streetNumber"
                        df.at[idx, "target"] = "streetNumber"
                    else:
                        # Otherwise, set 'target' to "streetName"
                        df.at[idx, "target"] = "streetName"

            # Return the modified DataFrame
            return df
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def reorder_and_filter(self) -> None:
        """
        Reorders the columns of the DataFrame, transforms the address column, filters specific rows,
        and performs various data processing steps including exploding, stacking, and renaming columns.

        This method processes the data in the following steps:
        - Reorders columns based on a predefined order
        - Stacks the DataFrame and reshapes it
        - Filters unwanted rows based on specific conditions
        - Explodes and renames address-related rows
        - Adds line numbers and merges with a count of total lines
        - Adds a new column 'AA' containing the value of 'text' from rows where 'target' is 'aa'
        - Filters the DataFrame to remove unnecessary rows

        Returns:
            None: This method modifies the class attribute `self.filtered_df` in place and does not return any value.
        """
        try:
            # Step 1: Reorder DataFrame columns based on a predefined order
            columns_initial_ordered = self.get_column_order(self.excel_df, self.columns_order)
            self.excel_df = self.excel_df[columns_initial_ordered]

            # Remove single/ double quotes from data
            self.excel_df = self.remove_quotes(self.excel_df)

            # Step 2: Transform the address column to remove or modify specific substrings
            self.excel_df = self.transform_address(self.excel_df)

            # Step 3: Stack the DataFrame to reshape it, converting columns into multi-level index
            stacked_df = self.excel_df.stack()

            # Step 4: Reset index after stacking to convert multi-level index into regular columns
            reset_stacked_df = stacked_df.reset_index()

            # Step 5: Rename columns for clarity after stacking
            # TODO: Ensure the column names are not hard-coded
            reset_stacked_df = reset_stacked_df.rename(
                columns={"level_0": "old_index", "level_1": "target", 0: "text"},
                errors="raise"
            )

            # Step 6: Filter out rows where "target" has specific unwanted values (e.g., "data_street", "data_number")
            column_to_filter = "target"
            values_to_filter = ["data_street", "data_number"]
            self.filtered_df = reset_stacked_df[~reset_stacked_df[column_to_filter].isin(values_to_filter)]

            # Step 7: Explode the "text" column, expanding list-like elements into separate rows if necessary
            self.filtered_df = self.explode_address(self.filtered_df)

            # Step 8: Rename rows for categorizing address components as "streetNumber" or "streetName"
            self.filtered_df = self.rename_address_rows(self.filtered_df)

            self.filtered_df["text"] = self.filtered_df["text"].apply(lambda x: str(x).strip())

            # Step 9: Add line numbers to the DataFrame
            self.filtered_df = self.add_line_numbers(self.filtered_df)

            # Step 10: Count the total lines for each "old_index" and merge the result into the DataFrame
            count_df = self.filtered_df.groupby(["old_index"]).count().iloc[:, 1].rename("total_lines")
            self.filtered_df = self.filtered_df.merge(count_df, how="left", left_on="old_index", right_index=True)

            # Step 11: Calculate the maximum line number for each group, excluding zero values
            self.filtered_df["total_lines"] = self.filtered_df.groupby("old_index")["line_number"].transform(
                lambda x: x[x != 0].max())

            # Step 12: Set "total_lines" to 0 for rows where "line_number" is zero
            self.filtered_df.loc[self.filtered_df["line_number"] == 0, "total_lines"] = 0

            # Step 13: Add a new column "AA" with the value of "text" from rows where "target" is "aa"
            self.filtered_df["AA"] = self.filtered_df.groupby("old_index")["text"].transform(
                lambda x: x[self.filtered_df["target"] == "aa"].values[0]
            )

            # Step 14: Keep only the relevant columns for the final DataFrame
            self.filtered_df = self.filtered_df[["old_index", "AA", "target", "text", "line_number", "total_lines"]]

            # Step 15: Remove rows where "target" is either "aa" or "full_address"
            column_to_filter = "target"
            values_to_filter = ["aa", "address"]
            self.filtered_df = self.filtered_df[~self.filtered_df[column_to_filter].isin(values_to_filter)]

            # Step 16: Corrects the values in the 'target' column of a DataFrame by standardizing them
            self.correct_target(self.filtered_df)
        except Exception as error:
            # Print and handle any error that occurs
            print(error)

    def generate_replacement_map(self, df: pd.DataFrame) -> dict:
        """
        Generates a replacement map for correcting values in the 'target' column of a DataFrame.
        The map is based on substrings before the dot in each 'target' value.

        E.g. columns {'toponym', 'toponym.1', 'toponym.2'} are mapped to correct column 'toponym'.

        Parameters:
            df (pd.DataFrame): The DataFrame containing the 'target' column that needs to be corrected.

        Returns:
            dict: A dictionary where each key is a substring before the dot in 'target' values,
                  and each value is a set of corresponding 'target' values.

        Raises:
            Exception: If an error occurs during the generation of the replacement map.
        """
        try:
            # Initialize an empty dictionary to store the replacement map
            replacement_map = {}

            # Retrieve distinct values from the 'target' column
            df_target_distinct_list = df['target'].unique()

            # Iterate through all distinct 'target' values
            for value in df_target_distinct_list:
                # Extract the substring before the dot from the current 'target' value
                prefix = value.split('.')[0]

                # Check if the prefix is already a key in the replacement map
                if prefix not in replacement_map:
                    # If the prefix is not in the map, initialize a new set with the current 'target' value
                    replacement_map[prefix] = {value}
                else:
                    # If the prefix exists, add the current 'target' value to the existing set
                    replacement_map[prefix].add(value)

            # Return the generated replacement map
            return replacement_map
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def correct_target(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Corrects the values in the 'target' column of a DataFrame by standardizing them.

        Parameters:
            df (pd.DataFrame): The DataFrame containing the 'target' column to be corrected.

        Returns:
            pd.DataFrame: The DataFrame with corrected 'target' values.

        Note:
            This function standardizes values in the 'target' column by removing suffixes
            (e.g., 'streetName.1' becomes 'streetName'). It checks against a replacement map
            generated from the DataFrame to ensure that correct values are retained.
        """
        try:
            # Generate a replacement map for the 'target' column
            replacements = self.generate_replacement_map(df)

            # Correct the 'target' column by standardizing values
            df["target"] = df["target"].apply(
                lambda target_value: target_value.split('.')[0] if target_value not in replacements else target_value
            )

            # Return the DataFrame with corrected 'target' values
            return df
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def write_in_csv(self) -> None:
        """
        Saves the filtered DataFrame to a CSV file.

        - The method constructs a file path using the attribute `self.output_csv_path`
          and saves the DataFrame as 'output.csv' in that directory.
        - It handles exceptions by printing any errors that occur during the process.
        """
        try:
            # Define the path for the output CSV file
            output_csv_path = os.path.join(self.output_csv_path, "output.csv")

            # Save the filtered DataFrame as a CSV file
            self.filtered_df.to_csv(output_csv_path, sep=";", index=False)
            print(f"Saved in CSV file 'output.csv' in {output_csv_path}.")
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def write_in_excel(self) -> None:
        """
        Writes the filtered DataFrame to a new sheet in the existing Excel file and also saves it as a CSV file.

        Returns:
            None: The method writes to an Excel file and a CSV file but does not return any value.

        Note:
            This function adds af new sheet named "output" to the existing Excel file, removes the sheet if it already exists,
            and saves the filtered DataFrame to this sheet. Additionally, it exports the DataFrame to a CSV file.
        """
        try:
            # Load the existing Excel workbook
            workbook = opyxl.load_workbook(filename=self.excel_file)

            # Define the name of the sheet to be created
            sheet_name = self.output_sheet

            # Check if the "output" sheet already exists in the workbook and delete it if present
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

            # Create a new sheet in the workbook with the name "output"
            ws = workbook.create_sheet(sheet_name)

            # Write the filtered DataFrame to the new Excel sheet, excluding the "old_index" column
            for row in dataframe_to_rows(self.filtered_df.loc[:, self.filtered_df.columns != "old_index"],
                                         index=False, header=True):
                ws.append(row)

            # Save the changes to the Excel file
            workbook.save(self.excel_file)
            print(f"Saved in new tab 'output' within '{self.excel_file}'.")
        except Exception as error:
            # Print and handle any error that occurs
            print(error)


def main(excel_file: str, columns_to_drop: list, output_csv_path: str, print_excel: bool = False) -> None:
    """
    Main function to process Excel data, correct values in the "target" column,
    and output the updated data to a new tab in the same Excel file and a CSV file.

    Parameters:
    - excel_file (str): The path to the Excel file to process.
    - columns_to_drop (list): A list of column names to be dropped from the Excel file.
    - output_csv_path (str): The directory path where the output CSV file will be saved.

    Returns:
    - None: The function processes the data and writes the results to files, but does not return any value.
    """
    try:
        # Print a message indicating the start of processing
        print(f"Processing for '{excel_file}'...")

        # Optional: Change the directory to the raw data directory (commented out)
        # os.chdir(r"../../data/raw")
        # excel_file = "input_output_template - Copy.xlsx"

        # Create an instance of the PreprocessingDataClass with the provided Excel file, columns to drop, and output path
        preprocessed_data = PreprocessingDataClass(excel_file, columns_to_drop, output_csv_path)

        # Call the method to drop specified columns
        preprocessed_data.drop_columns()

        # Call the method to construct a "rest" column in the Excel file
        preprocessed_data.build_rest()

        # Call the method to reorder and filter the DataFrame
        preprocessed_data.reorder_and_filter()

        # Call the method to write the processed DataFrame to .csv file
        preprocessed_data.write_in_csv()

        # Call the method to write the processed DataFrame to a new sheet in the Excel file and save as a CSV
        if print_excel:
            preprocessed_data.write_in_excel()
    except Exception as main_error:
        # Print any error encountered during the execution of the main function
        print(main_error)


if __name__ == '__main__':
    try:
        kwargs = {
            "excel_file": "input_output_template - Copy.xlsx",  # Set the Excel file with the input data
            "columns_to_drop": [],  # Set the columns to be dropped
            "output_csv_path": r"...",  # Set the path for where 'output.csv' will be saved
            "print_excel": False,  # Set whether to print or not the output to a new sheet in initial excel
        }
        main(**kwargs)
    except Exception as error:
        print(error)
