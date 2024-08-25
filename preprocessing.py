try:
    import os
    import pandas as pd
    import openpyxl as opyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    import re
except ImportError as import_err:
    print(import_err)

# Set pandas options
pd.set_option("display.max_columns", None)
pd.options.mode.copy_on_write = True


class PreprocessingDataClass:
    def __init__(self, excel_file: str, columns_to_drop: list[str], output_csv_path: str):
        self.excel_file = excel_file
        self.excel_df = pd.read_excel(self.excel_file, sheet_name='input', header=0)
        self.columns_to_drop = columns_to_drop
        self.output_csv_path = output_csv_path
        self.filtered_df = None
        self.columns_order = ["AA",
                              "source",
                              "streetName_prefix",
                              "Διεύθυνση",
                              "toponym",
                              "neighborhood",
                              "drop_field",
                              "klm",
                              "road",
                              "po_box",
                              "poi",
                              "data_city",
                              "data_zip",
                              "street",
                              "number",
                              "city",
                              "zip",
                              "type"]

    def get_column_order(self, df: pd.DataFrame, desired_order: list) -> list:
        """
        Reorders the columns of a DataFrame based on a desired order of prefixes.

        Parameters:
            df (pd.DataFrame): The DataFrame containing columns to reorder.
            desired_order (list): List of column prefixes in the desired order.

        Returns:
            list: List of columns ordered based on the desired prefixes.
        """
        try:
            # Initialize an empty list to hold the final order of columns
            final_order = []

            # Iterate over each prefix in the desired order list
            for column in desired_order:
                # Find columns in df that start with the current prefix and are not already in final_order
                matching_columns = [c for c in df.columns if c.startswith(column) and c not in final_order]

                # Add these matching columns to the final_order list
                final_order.extend(matching_columns)

            # Return the list of columns in the final order
            return final_order
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def drop_columns(self) -> None:
        """
        Drops specific columns from the DataFrame based on tags or a list of columns to drop.
        """
        try:
            # Check if any column is tagged as "drop_field"
            drop_fields = [column for column in self.excel_df.columns if "drop_field" in column]

            # Drop the columns that are tagged as "drop_field"
            self.excel_df = self.excel_df.drop(
                columns=self.excel_df.columns[self.excel_df.columns.map(lambda x: x in drop_fields)]
            )

            # Check if there are any columns specified to be dropped via the "columns_to_drop" attribute
            if self.columns_to_drop:
                # Verify that all columns to drop exist in the DataFrame
                if set(self.columns_to_drop) <= set(self.excel_df.columns.to_list()):
                    # Print the sorted list of columns that will be deleted
                    print(f"Deleting columns -> {sorted(self.columns_to_drop)}")

                    # Drop the columns specified in "columns_to_drop"
                    self.excel_df = self.excel_df.drop(
                        columns=self.excel_df.columns[
                            self.excel_df.columns.map(lambda x: x.split('.')[0] in self.columns_to_drop)]
                    )
                else:
                    # Print an error message if some columns in "columns_to_drop" do not exist in the DataFrame
                    print(
                        f"Input columns to drop: {set(self.columns_to_drop) - set(self.excel_df.columns.to_list())} are not valid."
                    )
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def transform_address(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Transforms the address column in the DataFrame by removing specific substrings found in other columns.

        Parameters:
            df (pd.DataFrame): The input DataFrame containing an address column and several other columns
                               with substrings that may need to be removed from the address.

        Returns:
            pd.DataFrame: The DataFrame with a transformed address column where specific substrings have been removed.
        """
        try:
            # Define the name of the address column to be transformed
            address_column = "Διεύθυνση"

            # Define a new column that will store the modified address
            new_address = "address_to_explode"

            # Define a temporary column to hold combined substrings to be removed from the address
            substrings_combined = "substrings_combined"

            # Copy the original address to the new address column
            df[new_address] = df[address_column]

            # Define a tuple of substring column names that need to be removed from the address
            exclude_substrings_column_tuple = (
                "po_box",
                "poi",
                "unit",
                "streetName_prefix",
                "toponym",
                "neighborhood",
                "drop_field",
                "klm"
            )

            # Iterate over each substring column in the tuple
            for substring_column in exclude_substrings_column_tuple:
                # Find columns in df that match the current substring column name
                matching_columns = [col for col in df.columns if substring_column in col]

                # If matching columns are found, combine their values into a single string per row
                if matching_columns:
                    df[substrings_combined] = df[matching_columns].apply(
                        lambda row: " ".join(row.dropna().astype(str).str.strip()), axis=1
                    )

                # Define a function to remove substrings from the address column
                def remove_substrings_from_address(row):
                    # Get the address from the new address column
                    address = row[new_address]

                    # Get the combined substrings to remove
                    substr_to_remove = row[substrings_combined]

                    # If the substring to remove is valid and exists in the address, remove it
                    if pd.notna(substr_to_remove) and substr_to_remove in address:
                        address = address.replace(substr_to_remove, '').strip()

                    # Return the cleaned address
                    return address

                # Apply the remove_substrings_from_address function to each row in the DataFrame
                df[new_address] = df.apply(remove_substrings_from_address, axis=1)

            # Drop the temporary substrings_combined column as it's no longer needed
            df = df.drop(substrings_combined, axis=1)

            # Return the transformed DataFrame
            return df
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def break_address(self, address: str) -> list[str]:
        """
        Breaks down an address string into a list of substrings, removing commas and empty strings.

        Parameters:
            address (str): The input address string that needs to be split into substrings.

        Returns:
            list[str]: A list of substrings obtained from the address, with commas removed and empty strings excluded.
        """
        try:
            # Replace commas in the address with an empty string, then split the address into substrings
            break_address = [substring for substring in address.replace(r',', '').split() if not substring == '']

            # Return the list of non-empty substrings
            return break_address
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def explode_address(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Splits address strings in the DataFrame into individual components and expands them into separate rows.

        Parameters:
            df (pd.DataFrame): The input DataFrame containing a "text" column that may contain address strings.

        Returns:
            pd.DataFrame: The DataFrame with address strings split into individual components and exploded into separate rows.
        """
        try:
            # Apply the break_address function to the "text" column if the "target" column is "address_to_explode"
            df["text"] = df.apply(
                lambda row: self.break_address(row["text"]) if row["target"] == "address_to_explode" else row["text"], axis=1
            )

            df = df.explode("text")
            return df
        except Exception as error:
            # Print and handle any error that occurs
            print(error)

    def rename_address_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Renames the 'target' column in the DataFrame based on the content of the 'text' column,
        distinguishing between street numbers and street names.

        Parameters:
            df (pd.DataFrame): The input DataFrame containing a 'target' column that identifies
                               address components and a 'text' column with the address data.

        Returns:
            pd.DataFrame: The DataFrame with updated 'target' values for rows where the 'target' was "address_to_explode".
        """
        try:
            # Reset the DataFrame's index and drop the old index column
            df = df.reset_index().drop(columns=["index"])
            df.fillna(value={"text": " "}, inplace=True)

            # Iterate over each row in the DataFrame
            for idx, row in df.iterrows():
                # Check if the 'target' column is "address_to_explode"
                if row["target"] == "address_to_explode":
                    # Check if 'text' is a number or contains minor non-numeric content like spaces or hyphens
                    if re.fullmatch(r"\d+[-\s]*\d*", row["text"]):
                        # If the text matches a street number pattern, set 'target' to "streetNumber"
                        df.at[idx, "target"] = "streetNumber"
                    else:
                        # Otherwise, set 'target' to "streetName"
                        df.at[idx, "target"] = "streetName"

            # Return the modified DataFrame
            return df
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def reorder_and_filter(self) -> None:
        """
        Reorders the columns of the DataFrame, transforms the address column, filters specific rows,
        and performs various data processing steps including exploding, stacking, and renaming columns.

        Parameters:
            None

        Returns:
            None: This method modifies the class attribute `self.filtered_df` in place and does not return any value.
        """
        try:
            # Reorder DataFrame columns based on a predefined order, moving "data_zip" and "data_city" to the end
            columns_initial_ordered = self.get_column_order(self.excel_df, self.columns_order)
            self.excel_df = self.excel_df[columns_initial_ordered]

            # Transform the address column to remove specific substrings
            self.excel_df = self.transform_address(self.excel_df)

            # Rename the address column from "Διεύθυνση" to "address"
            self.excel_df = self.excel_df.rename(columns={"Διεύθυνση": "address"})

            # Stack the DataFrame columns into indexes to reshape the DataFrame
            # TODO: Consider using pd.melt() instead of pd.stack() for this operation
            stacked_df = self.excel_df.stack()
            reset_stacked_df = stacked_df.reset_index()

            # Rename columns to reflect their new roles after stacking
            # TODO: Ensure the column names are not hard-coded
            reset_stacked_df = reset_stacked_df.rename(
                columns={"level_0": "old_index", "level_1": "target", 0: "text"},
                errors="raise"
            )

            # Filter out rows where "target" has specific unwanted values
            column_to_filter = "target"
            values_to_filter = ["data_street", "data_number"]
            self.filtered_df = reset_stacked_df[~reset_stacked_df[column_to_filter].isin(values_to_filter)]

            # Explode the "text" column, expanding list-like elements into separate rows
            self.filtered_df = self.explode_address(self.filtered_df)
            # self.filtered_df.drop(self.filtered_df[self.filtered_df["target"] == "Διεύθυνση"].index, inplace=True)

            # Merge the DataFrame with a count of total lines for each address
            count_df = self.filtered_df.groupby(["old_index"]).count().iloc[:, 1].rename("total_lines")
            self.filtered_df = self.filtered_df.merge(count_df, how="left", left_on="old_index", right_index=True)

            # Add a new column that counts from zero up to the total number of lines for each address
            self.filtered_df["line_number"] = self.filtered_df.groupby(["old_index", "total_lines"]).cumcount()

            # Correct wrong values in the "target" column using a custom method
            self.correct_target(self.filtered_df)

            # Reorder DataFrame columns, ensuring "total_lines" is moved to the end
            columns_to_reorder = ["total_lines"]
            column_list = [column for column in self.filtered_df.columns.to_list() if column not in columns_to_reorder]
            self.filtered_df = self.filtered_df.reindex(columns=column_list + columns_to_reorder)

            # Rename address-related rows to categorize them as either "streetNumber" or "streetName"
            self.filtered_df = self.rename_address_rows(self.filtered_df)
            # self.filtered_df = self.rename_address_rows(self.filtered_df).reset_index(drop=True)
        except Exception as error:
            # Print and handle any error that occurs
            print(error)

    def generate_replacement_map(self, df: pd.DataFrame) -> dict:
        """
        Generates a replacement map for correcting values in the 'target' column of a DataFrame.
        The map is based on substrings before the dot in each 'target' value.

        Parameters:
            df (pd.DataFrame): The DataFrame containing the 'target' column that needs to be corrected.

        Returns:
            dict: A dictionary where each key is a substring before the dot in 'target' values,
                  and each value is a set of corresponding 'target' values.

        Raises:
            Exception: If an error occurs during the generation of the replacement map.
        """
        try:
            # Initialize an empty dictionary to store the replacement map
            replacement_map = {}

            # Retrieve distinct values from the 'target' column
            df_target_distinct_list = df['target'].unique()

            # Iterate through all distinct 'target' values
            for value in df_target_distinct_list:
                # Extract the substring before the dot from the current 'target' value
                prefix = value.split('.')[0]

                # Check if the prefix is already a key in the replacement map
                if prefix not in replacement_map:
                    # If the prefix is not in the map, initialize a new set with the current 'target' value
                    replacement_map[prefix] = {value}
                else:
                    # If the prefix exists, add the current 'target' value to the existing set
                    replacement_map[prefix].add(value)

            # Return the generated replacement map
            return replacement_map
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def correct_target(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Corrects the values in the 'target' column of a DataFrame by standardizing them.

        Parameters:
            df (pd.DataFrame): The DataFrame containing the 'target' column to be corrected.

        Returns:
            pd.DataFrame: The DataFrame with corrected 'target' values.

        Note:
            This function standardizes values in the 'target' column by removing suffixes
            (e.g., 'streetName.1' becomes 'streetName'). It checks against a replacement map
            generated from the DataFrame to ensure that correct values are retained.
        """
        try:
            # Generate a replacement map for the 'target' column
            replacements = self.generate_replacement_map(df)

            # Correct the 'target' column by standardizing values
            df["target"] = df["target"].apply(
                lambda target_value: target_value.split('.')[0] if target_value not in replacements else target_value
            )

            # Return the DataFrame with corrected 'target' values
            return df
        except Exception as error:
            # Print any error encountered during execution
            print(error)

    def write_in_excel(self) -> None:
        """
        Writes the filtered DataFrame to a new sheet in the existing Excel file and also saves it as a CSV file.

        Parameters:
            None

        Returns:
            None: The method writes to an Excel file and a CSV file but does not return any value.

        Note:
            This function adds a new sheet named "output" to the existing Excel file, removes the sheet if it already exists,
            and saves the filtered DataFrame to this sheet. Additionally, it exports the DataFrame to a CSV file.
        """
        # Load the existing Excel workbook
        workbook = opyxl.load_workbook(filename=self.excel_file)

        # Define the name of the sheet to be created
        sheet_name = "output"

        # Check if the "output" sheet already exists in the workbook and delete it if present
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

        # Create a new sheet in the workbook with the name "output"
        ws = workbook.create_sheet(sheet_name)

        # Write the filtered DataFrame to the new Excel sheet, excluding the "old_index" column
        for row in dataframe_to_rows(self.filtered_df.loc[:, self.filtered_df.columns != "old_index"],
                                     index=False, header=True):
            ws.append(row)

        # Save the changes to the Excel file
        workbook.save(self.excel_file)
        print(f"Saved in new tab 'output' within '{self.excel_file}'.")

        # Define the path for the output CSV file
        output_csv_path = os.path.join(self.output_csv_path, "output.csv")

        # Save the filtered DataFrame as a CSV file
        self.filtered_df.to_csv(output_csv_path, index=False)
        print(f"Saved in CSV file 'output.csv' in {output_csv_path}.")


def main(excel_file: str, columns_to_drop: list, output_csv_path: str) -> None:
    """
    Main function to process Excel data, correct values in the "target" column,
    and output the updated data to a new tab in the same Excel file and a CSV file.

    Parameters:
    - excel_file (str): The path to the Excel file to process.
    - columns_to_drop (list): A list of column names to be dropped from the Excel file.
    - output_csv_path (str): The directory path where the output CSV file will be saved.

    Returns:
    - None: The function processes the data and writes the results to files, but does not return any value.
    """
    try:
        # Print a message indicating the start of processing
        print(f"Processing for '{excel_file}'...")

        # Optional: Change the directory to the raw data directory (commented out)
        # os.chdir(r"../../data/raw")
        # excel_file = "input_output_template - Copy.xlsx"

        # Create an instance of the PreprocessingDataClass with the provided Excel file, columns to drop, and output path
        preprocessed_data = PreprocessingDataClass(excel_file, columns_to_drop, output_csv_path)

        # Call the method to drop specified columns
        preprocessed_data.drop_columns()

        # Call the method to reorder and filter the DataFrame
        preprocessed_data.reorder_and_filter()

        # Call the method to write the processed DataFrame to a new sheet in the Excel file and save as a CSV
        preprocessed_data.write_in_excel()
    except Exception as main_error:
        # Print any error encountered during the execution of the main function
        print(main_error)


if __name__ == '__main__':
    try:
        kwargs = {
            "excel_file": "input_output_template - Copy.xlsx",  # Set the Excel file with the input data
            "columns_to_drop": [],  # Set the columns to be dropped
            "output_csv_path": r"..."  # Set the path for where 'output.csv' will be saved
        }
        main(**kwargs)
    except Exception as error:
        print(error)
